#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Image Generator for Novel-to-Webtoon Project
专门为网文改编项目设计的图像生成工具，集成ComfyUI API
"""

import json
import os
import time
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple

from comfyui_client import ComfyUIClient
from workflow_manager import WorkflowManager


class NovelWebtoonImageGenerator:
    def __init__(self, server_address: str = "127.0.0.1:8188",
                 workflows_dir: str = "comfyui-workflows",
                 output_dir: str = "generated_images"):
        """
        初始化网文改编图像生成器

        Args:
            server_address: ComfyUI服务器地址
            workflows_dir: 工作流文件目录
            output_dir: 输出图像目录
        """
        self.client = ComfyUIClient(server_address)
        self.workflow_manager = WorkflowManager(workflows_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # 默认工作流配置
        self.default_workflow = "default_workflow.json"

        # 角色一致性配置
        self.character_consistency = {
            '男主角': '融合吴彦祖的深邃眼神与胡歌的文雅气质，古装扮相英俊不凡',
            '女主角': '融合刘亦菲的仙气飘逸与金晨的灵动美感，古装造型典雅动人'
        }

    def process_excel_data(self, excel_path: str) -> List[Dict[str, Any]]:
        """
        从Excel文件读取图像生成任务

        Args:
            excel_path: Excel文件路径

        Returns:
            List[Dict]: 图像生成任务列表
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            ws = wb['完整制作流程']

            tasks = []

            for row_idx in range(2, ws.max_row + 1):  # 跳过标题行
                row_data = [cell.value for cell in ws[row_idx]]

                if len(row_data) >= 15 and row_data[14]:  # 确保有完整图像提示词
                    task = {
                        '集数': row_data[0],
                        '镜头编号': row_data[1],
                        '场景': row_data[6],
                        '镜头类型': row_data[7],
                        '人物提示词': row_data[12],
                        '背景提示词': row_data[13],
                        '完整提示词': row_data[14],
                        'output_filename': f"Episode-{row_data[0]:02d}-Shot-{row_data[1]:02d}"
                    }
                    tasks.append(task)

            print(f"从Excel加载了 {len(tasks)} 个图像生成任务")
            return tasks

        except Exception as e:
            print(f"处理Excel数据时出错: {e}")
            return []

    def enhance_prompt_with_consistency(self, original_prompt: str,
                                      character_hints: List[str] = None) -> str:
        """
        增强提示词以保持角色一致性

        Args:
            original_prompt: 原始提示词
            character_hints: 角色提示列表

        Returns:
            str: 增强后的提示词
        """
        enhanced_prompt = original_prompt

        # 如果提到角色，添加一致性描述
        if character_hints:
            for hint in character_hints:
                if hint in ['男主角', '叶凡']:
                    enhanced_prompt = enhanced_prompt.replace(
                        '男主角', f"男主角({self.character_consistency['男主角']})"
                    ).replace(
                        '叶凡', f"叶凡({self.character_consistency['男主角']})"
                    )
                elif hint in ['女主角', '苏灵溪']:
                    enhanced_prompt = enhanced_prompt.replace(
                        '女主角', f"女主角({self.character_consistency['女主角']})"
                    ).replace(
                        '苏灵溪', f"苏灵溪({self.character_consistency['女主角']})"
                    )

        return enhanced_prompt

    def generate_single_image(self, task: Dict[str, Any],
                            workflow_name: str = None,
                            custom_params: Dict[str, Any] = None) -> Tuple[bool, List[str]]:
        """
        生成单个镜头的图像

        Args:
            task: 图像生成任务
            workflow_name: 工作流名称
            custom_params: 自定义参数

        Returns:
            Tuple[bool, List[str]]: (是否成功, 生成的文件路径列表)
        """
        workflow_name = workflow_name or self.default_workflow
        prompt = task['完整提示词']

        # 增强提示词以保持角色一致性
        enhanced_prompt = self.enhance_prompt_with_consistency(
            prompt, ['男主角', '女主角', '叶凡', '苏灵溪']
        )

        # 构建参数修改字典
        param_modifications = {
            '6.text': enhanced_prompt,  # 正向提示词（假设节点6是CLIP Text Encode）
            '3.seed': int(time.time()) % 1000000,  # 随机种子
        }

        # 添加自定义参数
        if custom_params:
            param_modifications.update(custom_params)

        # 创建任务专用输出目录
        task_output_dir = self.output_dir / f"Episode_{task['集数']:02d}" / f"Shot_{task['镜头编号']:02d}"

        # 工作流路径
        workflow_path = self.workflow_manager.workflows_dir / workflow_name

        # 生成图像
        success, files = self.client.generate_images(
            str(workflow_path),
            param_modifications,
            str(task_output_dir)
        )

        if success and files:
            # 重命名文件以包含更多信息
            renamed_files = []
            for i, file_path in enumerate(files):
                file_ext = Path(file_path).suffix
                new_name = f"{task['output_filename']}-{i+1:02d}{file_ext}"
                new_path = task_output_dir / new_name

                os.rename(file_path, new_path)
                renamed_files.append(str(new_path))

            return True, renamed_files
        else:
            return False, []

    def batch_generate_from_excel(self, excel_path: str,
                                workflow_name: str = None,
                                max_concurrent: int = 1) -> Dict[str, Any]:
        """
        从Excel数据批量生成图像

        Args:
            excel_path: Excel文件路径
            workflow_name: 工作流名称
            max_concurrent: 最大并发数

        Returns:
            Dict: 生成结果统计
        """
        tasks = self.process_excel_data(excel_path)
        if not tasks:
            return {'success': False, 'message': '没有找到有效的生成任务'}

        workflow_name = workflow_name or self.default_workflow

        # 检查工作流文件是否存在
        workflow_path = self.workflow_manager.workflows_dir / workflow_name
        if not workflow_path.exists():
            return {
                'success': False,
                'message': f'工作流文件不存在: {workflow_path}',
                'available_workflows': self.workflow_manager.list_workflows()
            }

        results = {
            'total_tasks': len(tasks),
            'successful_tasks': 0,
            'failed_tasks': 0,
            'generated_files': [],
            'errors': []
        }

        print(f"开始批量图像生成，共 {len(tasks)} 个任务")

        for i, task in enumerate(tasks, 1):
            print(f"\n处理任务 {i}/{len(tasks)}: 第{task['集数']}集镜头{task['镜头编号']}")

            try:
                success, files = self.generate_single_image(task, workflow_name)

                if success:
                    results['successful_tasks'] += 1
                    results['generated_files'].extend(files)
                    print(f"✅ 生成成功，文件: {files}")
                else:
                    results['failed_tasks'] += 1
                    error_msg = f"第{task['集数']}集镜头{task['镜头编号']} 生成失败"
                    results['errors'].append(error_msg)
                    print(f"❌ {error_msg}")

            except Exception as e:
                results['failed_tasks'] += 1
                error_msg = f"第{task['集数']}集镜头{task['镜头编号']} 出现异常: {e}"
                results['errors'].append(error_msg)
                print(f"💥 {error_msg}")

            # 添加延迟避免服务器过载
            if i < len(tasks):
                time.sleep(2)

        # 生成结果报告
        success_rate = (results['successful_tasks'] / results['total_tasks']) * 100
        print(f"\n🎉 批量生成完成!")
        print(f"📊 成功率: {success_rate:.1f}% ({results['successful_tasks']}/{results['total_tasks']})")
        print(f"📁 生成文件数: {len(results['generated_files'])}")

        results['success'] = results['successful_tasks'] > 0
        return results

    def create_workflow_from_template(self, template_name: str = "basic_txt2img") -> str:
        """
        创建适合网文改编的基础工作流模板

        Args:
            template_name: 模板名称

        Returns:
            str: 创建的工作流文件路径
        """
        # 基础的文本到图像工作流模板
        basic_workflow = {
            "3": {
                "inputs": {
                    "seed": 156680208700286,
                    "steps": 20,
                    "cfg": 8,
                    "sampler_name": "euler",
                    "scheduler": "normal",
                    "denoise": 1,
                    "model": ["4", 0],
                    "positive": ["6", 0],
                    "negative": ["7", 0],
                    "latent_image": ["5", 0]
                },
                "class_type": "KSampler",
                "_meta": {"title": "KSampler"}
            },
            "4": {
                "inputs": {"ckpt_name": "v1-5-pruned-emaonly.ckpt"},
                "class_type": "CheckpointLoaderSimple",
                "_meta": {"title": "Load Checkpoint"}
            },
            "5": {
                "inputs": {"width": 512, "height": 768, "batch_size": 1},
                "class_type": "EmptyLatentImage",
                "_meta": {"title": "Empty Latent Image"}
            },
            "6": {
                "inputs": {"text": "beautiful chinese ancient style", "clip": ["4", 1]},
                "class_type": "CLIPTextEncode",
                "_meta": {"title": "CLIP Text Encode (Prompt)"}
            },
            "7": {
                "inputs": {
                    "text": "lowres, bad anatomy, bad hands, text, error, missing fingers, extra digit, fewer digits, cropped, worst quality, low quality, normal quality, jpeg artifacts, signature, watermark, username, blurry",
                    "clip": ["4", 1]
                },
                "class_type": "CLIPTextEncode",
                "_meta": {"title": "CLIP Text Encode (Negative)"}
            },
            "8": {
                "inputs": {"samples": ["3", 0], "vae": ["4", 2]},
                "class_type": "VAEDecode",
                "_meta": {"title": "VAE Decode"}
            },
            "9": {
                "inputs": {"filename_prefix": "ComfyUI", "images": ["8", 0]},
                "class_type": "SaveImage",
                "_meta": {"title": "Save Image"}
            }
        }

        # 保存工作流模板
        template_filename = f"{template_name}.json"
        template_path = self.workflow_manager.workflows_dir / template_filename

        with open(template_path, 'w', encoding='utf-8') as f:
            json.dump(basic_workflow, f, indent=2, ensure_ascii=False)

        print(f"工作流模板已创建: {template_path}")
        return str(template_path)


def main():
    """示例用法"""
    generator = NovelWebtoonImageGenerator()

    # 创建基础工作流模板（如果不存在）
    generator.create_workflow_from_template("default_workflow")

    # 示例：从Excel批量生成图像
    excel_path = "FullFlow-完整制作流程.xlsx"
    if os.path.exists(excel_path):
        results = generator.batch_generate_from_excel(excel_path)
        print(f"\n生成结果: {results}")
    else:
        print(f"Excel文件不存在: {excel_path}")


if __name__ == "__main__":
    main()