#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Generator for Novel-to-Webtoon Project
生成包含剧情、脚本、分镜、图像提示词的完整Excel文档
"""

import os
import re
import json
from typing import Dict, List, Tuple, Optional
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

class NovelProjectExcelGenerator:
    def __init__(self, project_root: str):
        """
        初始化Excel生成器

        Args:
            project_root: 项目根目录路径
        """
        self.project_root = Path(project_root)
        self.unified_data = []  # 统一的数据结构，记录完整流程

    def build_unified_data(self) -> List[Dict]:
        """
        构建统一的数据结构，将剧情、剧本、分镜、图像提示词关联在一起
        """
        # 读取所有数据
        plot_points = self.read_plot_breakdown()
        scripts = self.read_scripts()
        storyboards = self.read_storyboards()
        image_prompts = self.read_image_prompts()

        # 创建统一的数据结构
        unified_data = []

        # 以分镜为主线，关联其他数据
        for storyboard in storyboards:
            episode_num = storyboard['集数']
            shot_num = storyboard['镜头编号']

            # 查找对应的图像提示词
            matching_prompt = None
            for prompt in image_prompts:
                if prompt['集数'] == episode_num and prompt['镜头编号'] == shot_num:
                    matching_prompt = prompt
                    break

            # 查找对应的剧本
            matching_script = None
            for script in scripts:
                if script['集数'] == episode_num:
                    matching_script = script
                    break

            # 查找相关的剧情点
            related_plots = [plot for plot in plot_points if f"第{episode_num}集" in plot.get('分集建议', '')]

            unified_data.append({
                '集数': episode_num,
                '镜头编号': shot_num,
                '剧情描述': related_plots[0].get('完整描述', '') if related_plots else '',
                '冲突类型': related_plots[0].get('冲突类型', '') if related_plots else '',
                '情绪钩子': related_plots[0].get('情绪钩子', '') if related_plots else '',
                '剧本标题': matching_script.get('剧本标题', '') if matching_script else '',
                '场景': storyboard['场景'],
                '镜头类型': storyboard['镜头类型'],
                '拍摄角度': storyboard['拍摄角度'],
                '摄像机运动': storyboard['摄像机运动'],
                '对白': storyboard['对白'],
                '镜头描述': matching_prompt.get('镜头描述', '') if matching_prompt else '',
                '人物提示词': matching_prompt.get('人物提示词', '') if matching_prompt else '',
                '背景提示词': matching_prompt.get('背景提示词', '') if matching_prompt else '',
                '完整图像提示词': matching_prompt.get('完整合成提示词', '') if matching_prompt else ''
            })

        return unified_data

    def read_plot_breakdown(self) -> List[Dict]:
        """读取剧情拆解文件"""
        plot_file = self.project_root / 'plot-breakdown.md'
        if not plot_file.exists():
            return []

        with open(plot_file, 'r', encoding='utf-8') as f:
            content = f.read()

        # 解析剧情点 - 匹配【剧情X】格式
        plot_points = []
        pattern = r'【剧情(\d+)】(.+?)，(.+?)，(.+?)，第(\d+)集，状态：(.+?)(?=\n|$)'
        matches = re.findall(pattern, content, re.MULTILINE)

        for match in matches:
            point_num, description, conflict_vs, emotion_type, episode_num, status = match

            # 提取冲突描述（取vs前面的部分作为标题）
            title = conflict_vs.split('vs')[0].strip() if 'vs' in conflict_vs else conflict_vs
            conflict_type = conflict_vs if 'vs' in conflict_vs else ''

            plot_points.append({
                '剧情点编号': int(point_num),
                '标题': title,
                '完整描述': description,
                '冲突类型': conflict_type,
                '情绪钩子': emotion_type,
                '分集建议': f"第{episode_num}集",
                '状态': status
            })

        return plot_points

    def read_scripts(self) -> List[Dict]:
        """读取剧本文件"""
        scripts_dir = self.project_root / 'scripts'
        if not scripts_dir.exists():
            return []

        scripts = []
        for script_file in scripts_dir.glob('Episode-*.md'):
            episode_num = re.search(r'Episode-(\d+)', script_file.name)
            if not episode_num:
                continue

            with open(script_file, 'r', encoding='utf-8') as f:
                content = f.read()

            # 解析剧本内容 - 直接提取整个内容作为一个场景
            title_match = re.search(r'^第(\d+)集：(.+)', content, re.MULTILINE)
            if title_match:
                episode_title = title_match.group(2)
            else:
                episode_title = f"第{episode_num.group(1)}集"

            # 按段落分割内容，过滤空行
            paragraphs = [p.strip() for p in content.split('\n\n') if p.strip()]

            # 合并所有内容作为场景内容
            scene_content = '\n\n'.join(paragraphs[1:])  # 跳过标题行

            scripts.append({
                '集数': int(episode_num.group(1)),
                '剧本标题': episode_title,
                '场景标题': '完整剧本',
                '场景内容': scene_content[:1000] + '...' if len(scene_content) > 1000 else scene_content
            })

        return scripts

    def read_storyboards(self) -> List[Dict]:
        """读取分镜脚本文件"""
        storyboard_dir = self.project_root / 'storyboard'
        if not storyboard_dir.exists():
            return []

        storyboards = []
        for storyboard_file in storyboard_dir.glob('Episode-*-Storyboard.txt'):
            episode_num = re.search(r'Episode-(\d+)', storyboard_file.name)
            if not episode_num:
                continue

            with open(storyboard_file, 'r', encoding='utf-8') as f:
                content = f.read()

            # 解析分镜信息
            shots = re.findall(r'【镜头(\d+)-开始】\n(.*?)【镜头\1-结束】', content, re.DOTALL)

            for shot_num, shot_content in shots:
                # 解析镜头详细信息
                scene_match = re.search(r'场景：(.+)', shot_content)
                shot_type_match = re.search(r'镜头类型：(.+)', shot_content)
                angle_match = re.search(r'拍摄角度：(.+)', shot_content)
                movement_match = re.search(r'摄像机运动：(.+)', shot_content)
                dialogue_match = re.search(r'对白：(.+)', shot_content)

                storyboards.append({
                    '集数': int(episode_num.group(1)),
                    '镜头编号': int(shot_num),
                    '场景': scene_match.group(1) if scene_match else '',
                    '镜头类型': shot_type_match.group(1) if shot_type_match else '',
                    '拍摄角度': angle_match.group(1) if angle_match else '',
                    '摄像机运动': movement_match.group(1) if movement_match else '',
                    '对白': dialogue_match.group(1) if dialogue_match else '',
                    '完整内容': shot_content.strip()[:300] + '...' if len(shot_content.strip()) > 300 else shot_content.strip()
                })

        return storyboards

    def read_image_prompts(self) -> List[Dict]:
        """读取图像提示词文件"""
        prompts_dir = self.project_root / 'image-prompts'
        if not prompts_dir.exists():
            return []

        prompts = []

        # 读取自然语言中文提示词文件
        prompt_file = prompts_dir / 'Episode-01-02-Natural-Chinese-Prompts.txt'
        if prompt_file.exists():
            with open(prompt_file, 'r', encoding='utf-8') as f:
                content = f.read()

            # 解析镜头提示词
            shots = re.findall(r'===== 镜头(\d+)-开始（(.+?)）=====(.*?)(?====== 镜头|\Z)', content, re.DOTALL)
            current_episode = 1

            for shot_num, description, shot_content in shots:
                # 切换到第2集
                if int(shot_num) == 1 and len(prompts) > 0:
                    current_episode = 2

                # 解析人物、背景、完整提示词
                character_match = re.search(r'【人物提示词】\n(.+?)【背景提示词】', shot_content, re.DOTALL)
                background_match = re.search(r'【背景提示词】\n(.+?)【完整合成提示词】', shot_content, re.DOTALL)
                complete_match = re.search(r'【完整合成提示词】\n(.+?)(?====|$)', shot_content, re.DOTALL)

                prompts.append({
                    '集数': current_episode,
                    '镜头编号': int(shot_num),
                    '镜头描述': description,
                    '人物提示词': character_match.group(1).strip() if character_match else '',
                    '背景提示词': background_match.group(1).strip() if background_match else '',
                    '完整合成提示词': complete_match.group(1).strip() if complete_match else ''
                })

        return prompts

    def create_styled_workbook(self) -> Workbook:
        """创建带样式的工作簿"""
        wb = Workbook()

        # 定义样式
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')

        return wb, header_fill, header_font, border, center_alignment

    def create_unified_worksheet(self, wb: Workbook, styles: Tuple) -> None:
        """创建统一的完整流程工作表"""
        header_fill, header_font, border, center_alignment = styles

        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        ws = wb.create_sheet('完整制作流程')

        # 设置标题
        headers = [
            '集数', '镜头编号', '剧情描述', '冲突类型', '情绪钩子', '剧本标题',
            '场景', '镜头类型', '拍摄角度', '摄像机运动', '对白', '镜头描述',
            '人物提示词', '背景提示词', '完整图像提示词'
        ]
        ws.append(headers)

        # 应用标题样式
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment

        # 添加统一数据
        for data_row in self.unified_data:
            ws.append([
                data_row['集数'],
                data_row['镜头编号'],
                data_row['剧情描述'],
                data_row['冲突类型'],
                data_row['情绪钩子'],
                data_row['剧本标题'],
                data_row['场景'],
                data_row['镜头类型'],
                data_row['拍摄角度'],
                data_row['摄像机运动'],
                data_row['对白'],
                data_row['镜头描述'],
                data_row['人物提示词'][:100] + '...' if len(data_row['人物提示词']) > 100 else data_row['人物提示词'],
                data_row['背景提示词'][:100] + '...' if len(data_row['背景提示词']) > 100 else data_row['背景提示词'],
                data_row['完整图像提示词'][:200] + '...' if len(data_row['完整图像提示词']) > 200 else data_row['完整图像提示词']
            ])

        # 调整列宽
        column_widths = [8, 10, 30, 15, 12, 20, 15, 12, 12, 12, 25, 15, 40, 40, 60]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    def create_plot_worksheet(self, wb: Workbook, styles: Tuple) -> None:
        """创建剧情拆解工作表"""
        header_fill, header_font, border, center_alignment = styles

        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        ws = wb.create_sheet('剧情拆解')

        # 设置标题
        headers = ['剧情点编号', '标题', '完整描述', '冲突类型', '情绪钩子', '分集建议', '状态']
        ws.append(headers)

        # 应用标题样式
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment

        # 添加数据
        for plot_point in self.data['plot_breakdown']:
            ws.append([
                plot_point['剧情点编号'],
                plot_point['标题'],
                plot_point['完整描述'],
                plot_point['冲突类型'],
                plot_point['情绪钩子'],
                plot_point['分集建议'],
                plot_point['状态']
            ])

        # 调整列宽
        column_widths = [10, 20, 40, 20, 15, 12, 8]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    def create_scripts_worksheet(self, wb: Workbook, styles: Tuple) -> None:
        """创建剧本工作表"""
        header_fill, header_font, border, center_alignment = styles

        ws = wb.create_sheet('剧本内容')

        headers = ['集数', '剧本标题', '场景标题', '场景内容']
        ws.append(headers)

        # 应用标题样式
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment

        # 添加数据
        for script in self.data['scripts']:
            ws.append([
                script['集数'],
                script['剧本标题'],
                script['场景标题'],
                script['场景内容']
            ])

        # 调整列宽
        column_widths = [8, 25, 20, 60]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    def create_storyboards_worksheet(self, wb: Workbook, styles: Tuple) -> None:
        """创建分镜工作表"""
        header_fill, header_font, border, center_alignment = styles

        ws = wb.create_sheet('电影分镜')

        headers = ['集数', '镜头编号', '场景', '镜头类型', '拍摄角度', '摄像机运动', '对白', '完整内容']
        ws.append(headers)

        # 应用标题样式
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment

        # 添加数据
        for storyboard in self.data['storyboards']:
            ws.append([
                storyboard['集数'],
                storyboard['镜头编号'],
                storyboard['场景'],
                storyboard['镜头类型'],
                storyboard['拍摄角度'],
                storyboard['摄像机运动'],
                storyboard['对白'],
                storyboard['完整内容']
            ])

        # 调整列宽
        column_widths = [8, 10, 20, 12, 12, 12, 25, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    def create_prompts_worksheet(self, wb: Workbook, styles: Tuple) -> None:
        """创建图像提示词工作表"""
        header_fill, header_font, border, center_alignment = styles

        ws = wb.create_sheet('图像提示词')

        headers = ['集数', '镜头编号', '镜头描述', '人物提示词', '背景提示词', '完整合成提示词']
        ws.append(headers)

        # 应用标题样式
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment

        # 添加数据
        for prompt in self.data['image_prompts']:
            ws.append([
                prompt['集数'],
                prompt['镜头编号'],
                prompt['镜头描述'],
                prompt['人物提示词'],
                prompt['背景提示词'],
                prompt['完整合成提示词']
            ])

        # 调整列宽
        column_widths = [8, 10, 15, 50, 50, 80]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    def generate_excel(self, output_path: str) -> bool:
        """
        生成完整的Excel文档

        Args:
            output_path: 输出文件路径

        Returns:
            bool: 是否成功生成
        """
        try:
            # 构建统一数据结构
            print("正在构建统一数据结构...")
            self.unified_data = self.build_unified_data()

            # 创建工作簿和样式
            wb, header_fill, header_font, border, center_alignment = self.create_styled_workbook()
            styles = (header_fill, header_font, border, center_alignment)

            # 创建统一的完整流程工作表
            print("正在创建完整制作流程工作表...")
            self.create_unified_worksheet(wb, styles)

            # 保存文件
            wb.save(output_path)
            print(f"Excel文件已生成：{output_path}")

            # 打印统计信息
            print(f"\n=== 数据统计 ===")
            print(f"完整制作流程记录: {len(self.unified_data)}个")

            # 按集数统计
            episodes = {}
            for data in self.unified_data:
                ep = data['集数']
                if ep not in episodes:
                    episodes[ep] = 0
                episodes[ep] += 1

            for ep, count in sorted(episodes.items()):
                print(f"第{ep}集: {count}个镜头")

            return True

        except Exception as e:
            print(f"生成Excel时出错: {e}")
            return False

def main():
    """主函数"""
    # 项目根目录（当前目录）
    project_root = "D:/AI/novel-claude-project"

    # 输出文件路径
    output_path = "D:/AI/novel-claude-project/FullFlow-完整制作流程.xlsx"

    # 创建生成器并生成Excel
    generator = NovelProjectExcelGenerator(project_root)
    success = generator.generate_excel(output_path)

    if success:
        print("Excel文档生成成功！")
    else:
        print("Excel文档生成失败！")

if __name__ == "__main__":
    main()