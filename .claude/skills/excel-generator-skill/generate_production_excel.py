#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Production Excel Generator for Novel-to-Anime Project
生成包含剧本、图像提示词的完整制作Excel文档
"""

import os
import json
import re
from typing import Dict, List, Optional
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime

class ProductionExcelGenerator:
    def __init__(self, project_root: str):
        """
        初始化Excel生成器

        Args:
            project_root: 项目根目录路径
        """
        self.project_root = Path(project_root)

    def read_scripts(self) -> List[Dict]:
        """读取剧本文件"""
        scripts_dir = self.project_root / '01_scripts'
        scripts = []

        for episode_num in range(1, 6):  # 第1-5集
            script_file = scripts_dir / f'Episode-{episode_num:02d}.md'
            if not script_file.exists():
                continue

            with open(script_file, 'r', encoding='utf-8') as f:
                content = f.read()

            # 解析剧本内容
            # 提取场景和镜头
            scenes = []
            lines = content.split('\n')

            current_scene = None
            current_shots = []

            for line in lines:
                line = line.strip()

                # 识别场景标题
                if line.startswith('【场景'):
                    if current_scene:
                        current_scene['shots'] = current_shots
                        scenes.append(current_scene)
                    current_scene = {
                        'scene_title': line,
                        'shots': []
                    }
                    current_shots = []

                # 识别镜头
                elif line.startswith('【镜头'):
                    shot_info = {
                        'shot_number': line,
                        'content': []
                    }
                    current_shots.append(shot_info)
                elif current_shots and line:
                    current_shots[-1]['content'].append(line)

            # 添加最后一个场景
            if current_scene:
                current_scene['shots'] = current_shots
                scenes.append(current_scene)

            # 提取集数标题
            title_match = re.search(r'^第(\d+)集：(.+)', content, re.MULTILINE)
            episode_title = title_match.group(2) if title_match else f"第{episode_num}集"

            scripts.append({
                '集数': episode_num,
                '剧本标题': episode_title,
                '场景数': len(scenes),
                '场景': scenes,
                '原始内容': content
            })

        return scripts

    def read_image_prompts(self) -> List[Dict]:
        """读取图像提示词文件"""
        prompts_dir = self.project_root / '02_image_prompts'
        all_prompts = []

        for episode_num in range(1, 6):  # 第1-5集
            prompt_file = prompts_dir / f'Episode-{episode_num:02d}-Prompts.json'
            if not prompt_file.exists():
                continue

            with open(prompt_file, 'r', encoding='utf-8') as f:
                try:
                    prompts_data = json.load(f)
                except json.JSONDecodeError as e:
                    print(f"警告：无法解析 {prompt_file}，错误：{e}")
                    continue

            # 处理每个镜头的提示词
            for shot_data in prompts_data:
                # 从prompt中提取描述信息
                prompt_text = shot_data.get('prompt', '')

                # 尝试提取场景描述
                scene_desc = prompt_text[:100] + '...' if len(prompt_text) > 100 else prompt_text

                # 生成镜头描述
                shot_desc = f"镜头{shot_data.get('shot_number', '')}"

                all_prompts.append({
                    '集数': episode_num,
                    '镜头编号': f"镜头{shot_data.get('shot_number', '')}",
                    '镜头描述': shot_desc,
                    '人物提示词': ', '.join(shot_data.get('characters', [])),
                    '背景提示词': '',
                    '完整合成提示词': prompt_text,
                    '风格提示词': '',
                    '技术参数': '',
                    'has_character': shot_data.get('has_character', False),
                    'character_refs': ', '.join(shot_data.get('character_refs', []))
                })

        return all_prompts

    def create_styled_workbook(self) -> Workbook:
        """创建带样式的工作簿"""
        wb = Workbook()

        # 定义样式
        header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')

        return wb, header_fill, header_font, border, center_alignment

    def create_scripts_worksheet(self, wb: Workbook, styles: tuple, scripts: List[Dict]) -> None:
        """创建剧本工作表"""
        header_fill, header_font, border, center_alignment = styles

        # 删除默认工作表
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        ws = wb.create_sheet('剧本汇总', 0)

        # 设置标题
        headers = ['集数', '剧本标题', '场景编号', '镜头编号', '场景描述', '镜头内容']
        ws.append(headers)

        # 应用标题样式
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment

        # 添加数据
        for script in scripts:
            for scene_idx, scene in enumerate(script['场景'], 1):
                for shot in scene['shots']:
                    ws.append([
                        script['集数'],
                        script['剧本标题'],
                        f"场景{scene_idx}",
                        shot['shot_number'],
                        scene['scene_title'],
                        '\n'.join(shot['content'])
                    ])

        # 调整列宽
        column_widths = [8, 25, 12, 15, 30, 50]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    def create_prompts_worksheet(self, wb: Workbook, styles: tuple, prompts: List[Dict]) -> None:
        """创建图像提示词工作表"""
        header_fill, header_font, border, center_alignment = styles

        ws = wb.create_sheet('图像提示词汇总')

        # 设置标题
        headers = [
            '集数', '镜头编号', '镜头描述', '人物提示词',
            '背景提示词', '风格提示词', '技术参数', '完整合成提示词'
        ]
        ws.append(headers)

        # 应用标题样式
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment

        # 添加数据
        for prompt in prompts:
            ws.append([
                prompt['集数'],
                prompt['镜头编号'],
                prompt['镜头描述'],
                prompt['人物提示词'],
                prompt['背景提示词'],
                prompt['风格提示词'],
                prompt['技术参数'],
                prompt['完整合成提示词']
            ])

        # 调整列宽
        column_widths = [8, 12, 25, 40, 40, 25, 20, 60]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    def create_overview_worksheet(self, wb: Workbook, styles: tuple, scripts: List[Dict], prompts: List[Dict]) -> None:
        """创建项目概览工作表"""
        header_fill, header_font, border, center_alignment = styles

        ws = wb.create_sheet('项目概览')

        # 项目信息
        ws.append(['项目概览'])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([''])

        # 基本信息
        ws.append(['生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws.append(['项目路径', str(self.project_root)])
        ws.append(['处理集数', '第1-5集'])
        ws.append([''])

        # 统计信息
        ws.append(['统计信息'])
        ws.cell(row=7, column=1).font = Font(bold=True, size=12)
        ws.append([''])

        # 按集数统计
        ws.append(['集数', '剧本标题', '场景数', '镜头数(提示词)'])

        episode_stats = {}
        for prompt in prompts:
            ep = prompt['集数']
            if ep not in episode_stats:
                episode_stats[ep] = {'prompts': 0}
            episode_stats[ep]['prompts'] += 1

        for script in scripts:
            ep = script['集数']
            title = script['剧本标题']
            scenes = script['场景数']
            shots = episode_stats.get(ep, {}).get('prompts', 0)

            ws.append([f"第{ep}集", title, scenes, shots])

        # 总计
        ws.append([''])
        ws.append(['总计', '', f"{len(scripts)}集", f"{len(prompts)}个镜头"])

        # 样式调整
        for col in range(1, 5):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15

    def generate_excel(self, output_path: str) -> bool:
        """
        生成完整的Excel文档

        Args:
            output_path: 输出文件路径

        Returns:
            bool: 是否成功生成
        """
        try:
            print("正在读取剧本文件...")
            scripts = self.read_scripts()
            print(f"已读取 {len(scripts)} 个剧本文件")

            print("正在读取图像提示词文件...")
            prompts = self.read_image_prompts()
            print(f"已读取 {len(prompts)} 个图像提示词")

            print("正在创建Excel工作簿...")
            wb, header_fill, header_font, border, center_alignment = self.create_styled_workbook()
            styles = (header_fill, header_font, border, center_alignment)

            print("正在创建剧本汇总工作表...")
            self.create_scripts_worksheet(wb, styles, scripts)

            print("正在创建图像提示词汇总工作表...")
            self.create_prompts_worksheet(wb, styles, prompts)

            print("正在创建项目概览工作表...")
            self.create_overview_worksheet(wb, styles, scripts, prompts)

            # 保存文件
            print(f"正在保存Excel文件到: {output_path}")
            wb.save(output_path)

            # 打印统计信息
            print(f"\n=== 数据统计 ===")
            print(f"剧本文件: {len(scripts)} 个")
            print(f"图像提示词: {len(prompts)} 个")

            episodes = {}
            for prompt in prompts:
                ep = prompt['集数']
                if ep not in episodes:
                    episodes[ep] = 0
                episodes[ep] += 1

            print("\n各集镜头数:")
            for ep, count in sorted(episodes.items()):
                print(f"第{ep}集: {count}个镜头")

            return True

        except Exception as e:
            print(f"生成Excel时出错: {e}")
            import traceback
            traceback.print_exc()
            return False

def main():
    """主函数"""
    # 项目路径
    project_root = "D:/AI/novel-claude-project/outputs/run_20251220_195543_anime"

    # 输出文件路径
    output_dir = Path(project_root) / "06_final_excel"
    output_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    output_path = output_dir / f"Production-Data-20251220-195543.xlsx"

    # 创建生成器并生成Excel
    print("=== Production Excel Generator ===")
    print(f"项目路径: {project_root}")
    print(f"输出路径: {output_path}")
    print()

    generator = ProductionExcelGenerator(project_root)
    success = generator.generate_excel(str(output_path))

    if success:
        print(f"\n[成功] Excel文档生成成功！")
        print(f"[文件] 文件位置: {output_path}")
        print("\n[工作表] 包含工作表:")
        print("  - 剧本汇总：按集数和镜头整理的剧本内容")
        print("  - 图像提示词汇总：按集数和镜头整理的提示词")
        print("  - 项目概览：统计信息和章节概要")
    else:
        print("\n[失败] Excel文档生成失败！")

if __name__ == "__main__":
    main()